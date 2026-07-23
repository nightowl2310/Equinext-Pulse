#!/usr/bin/env python3
"""
net_report.py
=============

Builds the FULL detailed participant net-OI table for the whole archive and
writes it to a single CSV (net_report.csv). It is the exact table we worked out
by hand for a 5-day window -- Long, Short, daily net, a within-week running
cumulative, the same-weekday values from the last three weeks, and a WEEKLY NET
summary row per week -- now produced for every trading day, every actor, and all
three index instruments.

    net (daily)  = Long - Short                       (per participant, per day)
    weekly net   = sum of that ISO week's daily nets
    last week    = the SAME weekday's daily net, 7 days earlier
    2 / 3 wk ago = the same weekday, 14 / 21 days earlier
                   (blank if that day was a market holiday)

Nothing is stored in the database -- the net is a cheap derived number, so this
reads `participant_oi` and recomputes fresh every run (never stale). Read-only.

One CSV, columns:
  Category, Row Type, Date, Weekday, Client Type, Long, Short, Daily Net,
  Week Cumulative, Weekly Net, Last Week, 2 Wk Ago, 3 Wk Ago

Row Type is 'daily' for a normal day, or 'WEEKLY NET' for the per-week summary
row (on those rows Weekly Net / Last Week / 2 Wk Ago / 3 Wk Ago are WEEKLY nets).

    python net_report.py                      # full history -> net_report.csv
    python net_report.py --start 2026-04-01
    python net_report.py --preview 12         # also print the first rows
"""

import argparse
import csv
import sqlite3
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

import nse_oi_scraper as scraper          # reuse parse_date()

# instrument key -> (long column, short column, nice name, long header, short header)
CATEGORIES = {
    "futures": ("future_index_long", "future_index_short", "Index Futures",
                "Future Index Long", "Future Index Short"),
    "calls":   ("option_index_call_long", "option_index_call_short", "Index Calls",
                "Call Index Long", "Call Index Short"),
    "puts":    ("option_index_put_long", "option_index_put_short", "Index Puts",
                "Put Index Long", "Put Index Short"),
}
ACTORS = ["Client", "DII", "FII", "Pro"]   # order matches the hand-built sheet
TABLE = "participant_oi"
WEEKDAY = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
WEEKDAY_FULL = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

HEADER = ["Category", "Row Type", "Date", "Weekday", "Client Type",
          "Long", "Short", "Daily Net", "Week Cumulative", "Weekly Net",
          "Last Week", "2 Wk Ago", "3 Wk Ago"]


def iso_week(d: date) -> str:
    y, w, _ = d.isocalendar()
    return f"{y}-W{w:02d}"


def load(conn, start, end):
    """Read the four participants' rows into
    data[(date_iso, actor)][category] = (long, short), plus the sorted dates."""
    conn.row_factory = sqlite3.Row
    cols = []
    for long_col, short_col, *_ in CATEGORIES.values():
        cols += [long_col, short_col]
    actors_sql = ", ".join(f"'{a}'" for a in ACTORS)   # code constants, safe
    q = (f"SELECT date, participant_type, {', '.join(cols)} "
         f"FROM {TABLE} WHERE participant_type IN ({actors_sql})")
    params = []
    if start:
        q += " AND date >= ?"; params.append(start)
    if end:
        q += " AND date <= ?"; params.append(end)
    q += " ORDER BY date"

    data = {}
    dates = set()
    for r in conn.execute(q, params):
        dates.add(r["date"])
        cell = {}
        for key, (long_col, short_col, *_) in CATEGORIES.items():
            cell[key] = (r[long_col], r[short_col])
        data[(r["date"], r["participant_type"])] = cell
    return data, sorted(dates)


def build_rows(data, dates):
    """Produce every output row (daily + weekly-net) for all categories/actors."""
    def net(day_iso, actor, key):
        cell = data.get((day_iso, actor))
        if cell is None:
            return None
        long_v, short_v = cell[key]
        if long_v is None or short_v is None:
            return None
        return long_v - short_v

    def same_weekday(day_iso, weeks_back, actor, key):
        prior = (date.fromisoformat(day_iso) - timedelta(weeks=weeks_back)).isoformat()
        return net(prior, actor, key)      # None if that day has no data (holiday)

    rows = []
    for key, (_, _, cat_name, *_) in CATEGORIES.items():
        for actor in ACTORS:
            # group this actor's dates into ISO weeks, keep chronological order
            weeks = {}
            for d in dates:
                if (d, actor) in data:
                    weeks.setdefault(iso_week(date.fromisoformat(d)), []).append(d)
            ordered = sorted(weeks)
            weekly_net = {w: sum(net(d, actor, key) for d in weeks[w]) for w in ordered}

            for i, w in enumerate(ordered):
                days = weeks[w]
                cum = 0
                for j, d in enumerate(days):
                    dn = net(d, actor, key)
                    cum += dn
                    wd = WEEKDAY[date.fromisoformat(d).weekday()]
                    rows.append([
                        cat_name, "daily", d, wd, actor,
                        data[(d, actor)][key][0], data[(d, actor)][key][1], dn,
                        "" if j == 0 else cum,           # week cumulative (blank on week's 1st day)
                        "",                              # Weekly Net (only on the summary row)
                        _blank(same_weekday(d, 1, actor, key)),
                        _blank(same_weekday(d, 2, actor, key)),
                        _blank(same_weekday(d, 3, actor, key)),
                    ])
                # WEEKLY NET summary row: prior columns hold prior WEEKS' weekly nets
                span = f"{days[0][5:]}..{days[-1][5:]}"
                rows.append([
                    cat_name, "WEEKLY NET", f"{w} ({span})", "", actor,
                    "", "", "", "", weekly_net[w],
                    _blank(weekly_net[ordered[i - 1]] if i >= 1 else None),
                    _blank(weekly_net[ordered[i - 2]] if i >= 2 else None),
                    _blank(weekly_net[ordered[i - 3]] if i >= 3 else None),
                ])
    return rows


def _blank(v):
    return "" if v is None else v


# --------------------------------------------------------------------------- #
#  "FORM" report -- one file per instrument, laid out exactly like the         #
#  hand-built sheet: 5 trading days per actor, a running 5-day cumulative,     #
#  and the SAME weekday's Long/Short/Net one and two weeks back.               #
# --------------------------------------------------------------------------- #

# 8 columns; {long}/{short} are filled in per instrument.
#   daily net       = long - short  (that day)
#   cumulative net  = running total of daily net for that actor, carried across
#                     weeks (Monday continues from the previous market day)
#   delta           = today's daily net - the last market day's daily net
FORM_HEADER = ["Day", "Date", "Client Type", "{long}", "{short}",
               "daily net", "cumulative net", "delta"]


def _ddmmyyyy(day_iso: str) -> str:
    return date.fromisoformat(day_iso).strftime("%d-%m-%Y")


def build_form_rows(data, weeks_ordered, weeks_map, key):
    """Rows (list-of-lists, no header) for ONE instrument, matching the sheet,
    expanded to WEEKLY BLOCKS over the whole range (week-major layout).

    For every ISO week, oldest -> newest:
        a week-label separator row, then per actor (Client/DII/FII/Pro) that
        week's trading days in date order followed by a blank spacer.

    Columns per day:
      daily net       = long - short
      cumulative net  = that actor's running total of daily net, CONTINUOUS
                        across weeks (Monday = previous market day's total + today's net)
      delta           = today's daily net - the last market day's daily net
                        (day-over-day change; blank on the actor's very first day)
    """
    def cell(day_iso, actor):
        c = data.get((day_iso, actor))
        return None if c is None else c[key]        # (long, short) or None

    def net_of(day_iso, actor):
        c = cell(day_iso, actor)
        if c is None or c[0] is None or c[1] is None:
            return None
        return c[0] - c[1]

    # Pre-compute, per actor, the CONTINUOUS cumulative net and the day-over-day
    # delta over every trading day in the range -- nothing resets weekly, so
    # Monday is never blank (it carries on from the previous market day).
    all_days = sorted({d for w in weeks_ordered for d in weeks_map[w]})
    cumnet, delta = {}, {}
    for actor in ACTORS:
        run, prev = 0, None
        for d in all_days:
            dn = net_of(d, actor)
            if dn is None:
                continue
            run += dn
            cumnet[(actor, d)] = run
            delta[(actor, d)] = None if prev is None else dn - prev
            prev = dn

    ncol = len(FORM_HEADER)
    rows = []
    for w in weeks_ordered:
        wk_days = weeks_map[w]                       # already sorted, may be < 5 (partial week)
        span = f"{_ddmmyyyy(wk_days[0])} to {_ddmmyyyy(wk_days[-1])}"
        rows.append([f"Week {w} ({span})"] + [""] * (ncol - 1))   # week separator (col C empty)
        for actor in ACTORS:
            for d in wk_days:
                c = cell(d, actor)
                if c is None or c[0] is None or c[1] is None:
                    long_v = short_v = dn = ""
                else:
                    long_v, short_v = c
                    dn = long_v - short_v
                rows.append([
                    WEEKDAY_FULL[date.fromisoformat(d).weekday()], _ddmmyyyy(d), actor,
                    long_v, short_v, dn,
                    _blank(cumnet.get((actor, d))),   # continuous cumulative net
                    _blank(delta.get((actor, d))),    # day-over-day delta
                ])
            rows.append([""] * ncol)                # blank spacer between actors
    return rows


def write_form_xlsx(rows, header, sheet_name, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    head_fill = PatternFill("solid", fgColor="1F4E78")
    head_font = Font(bold=True, color="FFFFFF")
    week_fill = PatternFill("solid", fgColor="D9E1F2")
    week_font = Font(bold=True, color="1F4E78")
    ws.append(header)
    for c in ws[1]:
        c.fill, c.font = head_fill, head_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in rows:
        ws.append(r)
        # a week-separator row has col A filled but no Client Type (col C) -> shade it
        if r[0] and not r[2]:
            for c in ws[ws.max_row]:
                c.fill, c.font = week_fill, week_font

    # numeric formatting + column widths
    for col in range(4, len(header) + 1):        # numeric columns start at "long"
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, int):
                cell.number_format = "#,##0"
    widths = [12, 12, 13, 16, 16, 14, 16, 14]   # Day, Date, Client Type, long, short, daily net, cumulative net, delta
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    wb.save(out_path)
    return out_path


def build_form(conn, start, end, out_prefix):
    """Group the whole range into ISO weeks, then write one xlsx per instrument."""
    _, range_dates = load(conn, start, end)      # trading days actually in [start, end]
    if not range_dates:
        return [], []

    # same ISO-week grouping the CSV path uses, so the two can't diverge
    weeks_map = {}
    for d in range_dates:
        weeks_map.setdefault(iso_week(date.fromisoformat(d)), []).append(d)
    weeks_ordered = sorted(weeks_map)
    for w in weeks_ordered:
        weeks_map[w].sort()

    # load a -21d buffer before the first day so early weeks' look-backs resolve
    hist_start = (date.fromisoformat(range_dates[0]) - timedelta(days=21)).isoformat()
    data, _ = load(conn, hist_start, range_dates[-1])

    outputs = []
    for key, (_, _, cat_name, long_h, short_h) in CATEGORIES.items():
        header = [h.replace("{long}", long_h).replace("{short}", short_h) for h in FORM_HEADER]
        rows = build_form_rows(data, weeks_ordered, weeks_map, key)
        out = f"{out_prefix}_{key}.xlsx"
        write_form_xlsx(rows, header, cat_name, out)
        outputs.append((cat_name, out))
    return outputs, (range_dates, weeks_ordered)


def write_csv(rows, out_path):
    with open(out_path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(HEADER)
        w.writerows(rows)
    return out_path


def main():
    p = argparse.ArgumentParser(
        description="Full detailed participant net-OI table (one CSV) for the 3 index instruments.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=("Examples:\n"
                "  python net_report.py\n"
                "  python net_report.py --start 2026-04-01 --end 2026-07-20\n"
                "  python net_report.py --preview 12\n"))
    p.add_argument("--db", default="nse_data.db", help="Database file (default: nse_data.db).")
    p.add_argument("--start", type=scraper.parse_date, help="First date YYYY-MM-DD (default: everything).")
    p.add_argument("--end", type=scraper.parse_date, help="Last date YYYY-MM-DD (default: everything).")
    p.add_argument("--out", default="net_report.csv", help="Output CSV (default: net_report.csv).")
    p.add_argument("--preview", type=int, default=0, metavar="N", help="Also print the first N rows.")
    p.add_argument("--form", action="store_true",
                   help="Write the 'form' workbooks (one xlsx per instrument): weekly blocks over the "
                        "whole range, Client/DII/FII/Pro per week, with the 5-day cumulative + last/second-last week.")
    p.add_argument("--form-prefix", default="net_report_form",
                   help="Form mode: output filename prefix (default: net_report_form -> net_report_form_futures.xlsx, ...).")
    args = p.parse_args()

    start = args.start.isoformat() if args.start else None
    end = args.end.isoformat() if args.end else None

    if args.form:
        conn = sqlite3.connect(args.db)
        try:
            outputs, meta = build_form(conn, start, end, args.form_prefix)
        finally:
            conn.close()
        if not outputs:
            print("ERROR: no data for that range -- is the database loaded?"); sys.exit(1)
        range_dates, weeks_ordered = meta
        print(f"Range   : {range_dates[0]} -> {range_dates[-1]}  "
              f"({len(range_dates)} trading days, {len(weeks_ordered)} weeks)")
        print(f"Layout  : week-major -- each week has Client/DII/FII/Pro")
        print(f"Columns : daily net, cumulative net (continuous), delta (day-over-day)")
        for cat_name, out in outputs:
            print(f"Wrote   : {cat_name:14s} -> {Path(out).resolve()}")
        return

    conn = sqlite3.connect(args.db)
    try:
        data, dates = load(conn, start, end)
    finally:
        conn.close()
    if not dates:
        print("ERROR: no data for that range -- is the database loaded?"); sys.exit(1)

    rows = build_rows(data, dates)
    out = write_csv(rows, args.out)

    n_daily = sum(1 for r in rows if r[1] == "daily")
    n_weekly = sum(1 for r in rows if r[1] == "WEEKLY NET")
    print(f"Range   : {dates[0]} -> {dates[-1]}  ({len(dates)} trading days)")
    print(f"Wrote   : {Path(out).resolve()}")
    print(f"Rows    : {len(rows)}  ({n_daily} daily + {n_weekly} weekly-net) "
          f"across {len(CATEGORIES)} instruments x {len(ACTORS)} actors")
    print(f"Columns : {', '.join(HEADER)}")

    if args.preview:
        print("\n--- preview ---")
        print(",".join(HEADER))
        for r in rows[:args.preview]:
            print(",".join(str(x) for x in r))


if __name__ == "__main__":
    main()
